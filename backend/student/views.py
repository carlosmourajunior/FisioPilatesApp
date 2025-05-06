from rest_framework import viewsets, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Student
from .serializers import StudentSerializer
from physiotherapist.models import Physiotherapist
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.utils.timezone import make_aware
from datetime import datetime
import io

class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = Student.objects.all()
        if not self.request.user.is_staff:
            # If user is a physiotherapist, only show their students
            try:
                physiotherapist = self.request.user.physiotherapist
                queryset = queryset.filter(physiotherapist=physiotherapist)
            except:
                queryset = Student.objects.none()
                
        active = self.request.query_params.get('active', None)
        if active is not None:
            queryset = queryset.filter(active=active.lower() == 'true')
        return queryset

    def perform_create(self, serializer):
        if not self.request.user.is_staff:
            # Se o usuário é um fisioterapeuta, atribuir automaticamente
            try:
                physiotherapist = self.request.user.physiotherapist
                serializer.save(physiotherapist=physiotherapist)
            except:
                serializer.save()
        else:
            # Se é admin, usa o fisioterapeuta selecionado no frontend ou None
            serializer.save()

    def perform_update(self, serializer):
        if not self.request.user.is_staff:
            # Se o usuário é um fisioterapeuta, manter ele como responsável
            try:
                physiotherapist = self.request.user.physiotherapist
                serializer.save(physiotherapist=physiotherapist)
            except:
                serializer.save()
        else:
            # Se é admin, usa o fisioterapeuta selecionado no frontend ou mantém o atual
            serializer.save()
    
    @action(detail=False, methods=['get'])
    def template(self, request):
        """Download a template for student import"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Importar Alunos"
        
        # Define headers
        headers = [
            'Nome*', 'Email', 'Telefone', 'Data de Nascimento', 'Observações',
            'Tipo de Pagamento (PRE/POS)', 'Ativo (SIM/NAO)', 'Comissão (%)'
        ]
        
        # Set column headers and style
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Set column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
            
        # Add example row
        example = [
            'João Silva', 'joao@email.com', '(11) 99999-9999', '1990-01-01',
            'Observações do aluno', 'PRE', 'SIM', '50'
        ]
        for col, value in enumerate(example, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Create the response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=template_importacao_alunos.xlsx'
        
        return response
        
    @action(detail=False, methods=['post'])
    def import_students(self, request):
        """Import students from Excel file"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Nenhum arquivo foi enviado'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        excel_file = request.FILES['file']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Skip header row
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            created_students = []
            errors = []
            
            for row_num, row in enumerate(rows, 2):
                try:
                    if not row[0]:  # Skip empty rows
                        continue
                        
                    student_data = {
                        'name': row[0],
                        'email': row[1] if row[1] else None,
                        'phone': row[2] if row[2] else None,
                        'date_of_birth': row[3] if row[3] else None,
                        'notes': row[4] if row[4] else None,
                        'payment_type': 'PRE' if row[5] in ['PRE', 'pre'] else 'POS',
                        'active': row[6].upper() == 'SIM' if row[6] else True,
                        'commission': float(row[7]) if row[7] else 50.0
                    }
                    
                    # Validate required fields
                    if not student_data['name']:
                        raise ValueError('Nome é obrigatório')
                        
                    # Create student
                    if not self.request.user.is_staff:
                        student_data['physiotherapist'] = self.request.user.physiotherapist.id
                        
                    serializer = self.get_serializer(data=student_data)
                    if serializer.is_valid():
                        serializer.save()
                        created_students.append(student_data['name'])
                    else:
                        errors.append({
                            'row': row_num,
                            'name': student_data['name'],
                            'errors': serializer.errors
                        })
                        
                except Exception as e:
                    errors.append({
                        'row': row_num,
                        'name': row[0] if row[0] else 'Linha vazia',
                        'errors': str(e)
                    })
                    
            return Response({
                'success': True,
                'created': len(created_students),
                'created_students': created_students,
                'errors': errors
            })
            
        except Exception as e:
            return Response(
                {'error': f'Erro ao processar arquivo: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Template Alunos"
        
        # Cria uma nova aba para listar as modalidades
        modalities_sheet = workbook.create_sheet(title="Modalidades")
        modalities_sheet.append(['ID', 'Nome da Modalidade'])
        
        # Busca todas as modalidades e adiciona na aba
        from modality.models import Modality
        modalities = Modality.objects.all()
        for modality in modalities:
            modalities_sheet.append([modality.id, modality.name])
            
        # Ajusta largura das colunas
        modalities_sheet.column_dimensions['A'].width = 10
        modalities_sheet.column_dimensions['B'].width = 30

        # Volta para a primeira aba
        sheet = workbook.active
        
        # Define headers
        headers = [
            'Nome',
            'Email',
            'Telefone',
            'Data de Nascimento (DD/MM/AAAA)',
            'Modalidade (ID)',
            'Data de Registro (DD/MM/AAAA)',
            'Comissão (%)',
            'Dias da Semana (SEG,TER,QUA,QUI,SEX,SAB)',
            'Horário (HH:MM)',
            'Ativo (Sim/Não)',
            'Observações'
        ]
        
        # Adiciona nota sobre a aba de modalidades
        sheet['A2'] = "IMPORTANTE: Consulte a aba 'Modalidades' para ver os IDs disponíveis"
        sheet['A2'].font = Font(bold=True, color="FF0000")

        # Estilo para headers
        header_font = Font(bold=True)

        # Adiciona headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            sheet.column_dimensions[get_column_letter(col)].width = 20

        # Cria o response com o arquivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=template_alunos.xlsx'

        # Salva o workbook no response
        workbook.save(response)
        return response

    @action(detail=False, methods=['post'])
    def upload(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'Nenhum arquivo enviado'}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        if not file.name.endswith('.xlsx'):
            return Response({'error': 'Arquivo deve ser .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Carrega o arquivo Excel
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active            # Pula a primeira linha (cabeçalho)
            for row in sheet.iter_rows(min_row=2, values_only=True):                
                if not row[0]:  # Se o nome está vazio, pula a linha
                    continue

                # Processa a data de nascimento
                birth_date = None
                if row[3]:
                    if isinstance(row[3], datetime):
                        birth_date = row[3].strftime('%Y-%m-%d')
                    else:
                        try:
                            birth_date = datetime.strptime(row[3], '%d/%m/%Y').strftime('%Y-%m-%d')
                        except ValueError:
                            return Response(
                                {'error': f'Formato de data de nascimento inválido para o aluno {row[0]}. Use o formato DD/MM/AAAA'},
                                status=status.HTTP_400_BAD_REQUEST
                            )

                # Processa a data de registro
                registration_date = None
                if row[5]:
                    if isinstance(row[5], datetime):
                        registration_date = row[5].strftime('%Y-%m-%d')
                    else:
                        try:
                            registration_date = datetime.strptime(row[5], '%d/%m/%Y').strftime('%Y-%m-%d')
                        except ValueError:
                            return Response(
                                {'error': f'Formato de data de registro inválido para o aluno {row[0]}. Use o formato DD/MM/AAAA'},
                                status=status.HTTP_400_BAD_REQUEST
                            )
                
                # Processa os dias da semana
                weekdays = row[7].split(',') if row[7] else []
                  # Trata o ID da modalidade
                try:
                    modality_id = int(row[4])
                except (ValueError, TypeError):
                    return Response(
                        {'error': f'ID da modalidade inválido na linha {row[0]}. Deve ser um número inteiro.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Verifica se a modalidade existe
                from modality.models import Modality
                if not Modality.objects.filter(id=modality_id).exists():
                    return Response(
                        {'error': f'Modalidade com ID {modality_id} não encontrada para o aluno {row[0]}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )                # Converte os dias da semana em uma lista
                weekdays_map = {
                    'SEG': 0, 'TER': 1, 'QUA': 2, 
                    'QUI': 3, 'SEX': 4, 'SAB': 5, 'DOM': 6
                }
                weekday_list = []
                if row[7]:  # Dias da semana
                    weekday_list = [weekdays_map[day.strip().upper()] for day in row[7].split(',')]                # Converte o horário para inteiro (hora)
                hour = None
                if row[8]:  # Horário
                    try:
                        if isinstance(row[8], datetime):
                            # Se for datetime, pega a hora
                            hour = row[8].hour
                        elif isinstance(row[8], str):
                            # Se for string, extrai a hora
                            hour = int(row[8].split(':')[0])
                        else:
                            # Para outros tipos (como time)
                            hour = row[8].hour if hasattr(row[8], 'hour') else int(str(row[8]).split(':')[0])
                    except (ValueError, IndexError, AttributeError):
                        return Response(
                            {'error': f'Formato de horário inválido para o aluno {row[0]}. Use o formato HH:MM'},
                            status=status.HTTP_400_BAD_REQUEST
                        )# Adiciona o fisioterapeuta atual se não for admin
                if not self.request.user.is_staff:
                    physiotherapist_id = self.request.user.physiotherapist.id
                else:
                    physiotherapist_id = None

                # Cria o aluno
                student_data = {
                    'name': row[0],
                    'email': row[1],
                    'phone': row[2],
                    'date_of_birth': birth_date,
                    'modality': modality_id,
                    'registration_date': registration_date,
                    'commission': float(row[6]) if row[6] else 50.0,
                    'active': row[9].lower() == 'sim',
                    'notes': row[10],
                    'physiotherapist': physiotherapist_id
                }                
                serializer = self.get_serializer(data=student_data)
                if serializer.is_valid():
                    # Salva o aluno e pega a instância criada
                    serializer.save()
                    student = serializer.instance
                    
                    # Cria os horários para o aluno
                    if weekday_list and hour is not None:
                        from schedule.models import StudentSchedule
                        from schedule.serializers import StudentScheduleSerializer
                        
                        for weekday in weekday_list:
                            schedule_data = {
                                'student': student.id,
                                'weekday': weekday,
                                'hour': hour
                            }
                            schedule_serializer = StudentScheduleSerializer(data=schedule_data)
                            if schedule_serializer.is_valid():
                                schedule_serializer.save()
                            else:
                                return Response(
                                    {'error': f'Erro ao criar horário para o aluno {row[0]}: {schedule_serializer.errors}'},
                                    status=status.HTTP_400_BAD_REQUEST
                                )
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({'message': 'Alunos importados com sucesso'}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
